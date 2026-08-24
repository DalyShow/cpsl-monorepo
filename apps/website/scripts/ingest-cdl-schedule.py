#!/usr/bin/env python3
"""
Ingest the CDL Fall 2026 schedule from an Excel file into normalised JSON.

Reads:
  ~/Desktop/CDL - Schedules and Contact Sheet Fall 2026.xlsx  (default)
  OR the path passed as arg 1

Writes:
  apps/website/data/cdl-fall-2026.json
      { generatedAt, source, clubs[], matches[] }
  apps/website/data/cdl-fall-2026.warnings.json
      { row: {…}, warnings: [...] }

Design notes:
  - Reads the "Fall Schedule" sheet only. "DivisionName" = age group (per the
    club president). We split the age from any trailing M/A/B suffix.
  - Every team string is matched against a canonical CDL_CLUBS list of
    { name, aliases[] }. First match wins. If none match, we drop into
    a fallback ("Unknown Club") and log a warning.
  - Time cells arrive as either datetime.time (72% of rows) or strings
    like "10.30 AM" / "9.00 am" — parse_time handles both.
  - Kickoff is emitted as a naive ISO string (YYYY-MM-DDTHH:MM:SS) with no
    timezone offset. The site renders in local time, so this is safe.
"""

from __future__ import annotations

import datetime as dt
import json
import os
import re
import sys
from pathlib import Path

import openpyxl

# ─── Paths ──────────────────────────────────────────────────────────────────

REPO_ROOT = Path(__file__).resolve().parents[3]        # cpsl-monorepo/
APP_DIR   = REPO_ROOT / "apps" / "website"
DATA_DIR  = APP_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

BOYS_SOURCE          = Path.home() / "Desktop" / "CDL - Schedules and Contact Sheet Fall 2026.xlsx"
BOYS_SNAPSHOT_JSON   = DATA_DIR / "cdl-fall-2026-boys-snapshot.json"
GIRLS_SOURCE         = Path.home() / "Desktop" / "CDL - Schedules and Contact Sheet Fall 2026 (This is going to CDL_CPSL Web person.xlsx"

# One entry per feed. `kind` = "xlsx" or "snapshot" (JSON already parsed).
# `snapshot_fallback` is used when the live xlsx isn't on disk (see boys).
FEEDS = [
    {
        "kind":             "xlsx",
        "path":             BOYS_SOURCE,
        "tab":              "Fall Schedule",
        "gender":           "M",
        "has_location":     True,
        "snapshot_fallback": BOYS_SNAPSHOT_JSON,  # used when the .xlsx is missing
    },
    {
        "kind":             "xlsx",
        "path":             GIRLS_SOURCE,
        "tab":              "Girls CDL Fall Schedule",
        "gender":           "G",
        "has_location":     False,
    },
]

OUT_MATCHES  = DATA_DIR / "cdl-fall-2026.json"
OUT_WARNINGS = DATA_DIR / "cdl-fall-2026.warnings.json"

# ─── Canonical clubs ───────────────────────────────────────────────────────
# `aliases` are case-insensitive substrings that identify the club within a
# team-name string. The longest, most specific alias should come first — the
# matcher scores by longest hit so "NC Fusion" beats "NC" beats "".

CDL_CLUBS = [
    {
        "id":       "csa",
        "name":     "CSA",
        "short":    "CSA",
        "aliases":  ["charlotte soccer academy", "csa clt", "csa nh", "csa north", "csa charlotte", "csa"],
        # "csa" alone is too greedy — matches "loco-acsandhills.png" as a
        # substring. Anchor on filename structure so we only hit real CSA crests.
        "sanityNameHints": ["charlotte soccer academy", "logo-csa", "-csa.", "/csa"],
        # Tokens safe to strip from the team string when deriving the team
        # label — must identify the CLUB and nothing else (no location or
        # colour suffixes like "ncf ws" or "wake fc north").
        "label_strip":     ["charlotte soccer academy", "csa"],
    },
    {
        "id":       "cisc",
        "name":     "CISC",
        "short":    "CISC",
        "aliases":  ["cisc north", "cisc south", "cisc east", "cisc", "pre mls", "independence"],
        "sanityNameHints": ["independence", "cisc", "charlotte independence"],
        "label_strip":     ["charlotte independence", "cisc"],
    },
    {
        "id":       "nc-fusion",
        "name":     "NC Fusion",
        "short":    "NCF",
        "aliases":  ["north carolina fusion", "nc fusion", "ncf pre", "ncf ws", "ncf gso", "ws red", "gso red", "nc fus", "fusion", " ncf "],
        "sanityNameHints": ["fusion", "nc fusion", "north carolina fusion", "ncf"],
        "label_strip":     ["north carolina fusion", "nc fusion", "fusion", "ncf"],
    },
    {
        "id":       "nc-fc-youth",
        "name":     "NC FC Youth",
        "short":    "NCFC",
        "aliases":  ["north carolina fc youth", "ncfc north", "ncfc south", "ncfc", "nc fc"],
        "sanityNameHints": ["ncfc", "north carolina fc", "nc fc"],
        "label_strip":     ["north carolina fc youth", "nc fc youth", "ncfc", "nc fc"],
    },
    {
        "id":       "wilmington-hammerheads",
        "name":     "Wilmington Hammerheads Youth",
        "short":    "WHYFC",
        "aliases":  ["wilmington hammerheads youth", "wilmington hammerheads", "wilmington hh", "hammerheads", "whyfc"],
        "sanityNameHints": ["hammerheads", "wilmington", "whyfc"],
        "label_strip":     ["wilmington hammerheads youth", "wilmington hammerheads", "wilmington hh", "hammerheads", "whyfc"],
    },
    {
        "id":       "highland-fc",
        "name":     "Highland FC",
        "short":    "HFC",
        "aliases":  ["highland fc", "highland", "hfc"],
        "sanityNameHints": ["highland", "hfc"],
        "label_strip":     ["highland fc", "highland", "hfc"],
    },
    {
        "id":       "sc-surf",
        "name":     "SC Surf",
        "short":    "SCS",
        "aliases":  ["sc surf", "surf"],
        "sanityNameHints": ["sc surf", "surf"],
        "label_strip":     ["sc surf"],
    },
    {
        "id":       "wake-fc",
        "name":     "Wake FC",
        "short":    "WFC",
        "aliases":  ["wake fc north", "wake fc south", "wake fc", "wake", "wfc pa", "wfc", "w fc pag", "w fc"],
        "sanityNameHints": ["wake", "wfc"],
        "label_strip":     ["wake fc", "wake", "wfc", "w fc"],
    },
    {
        "id":       "carolina-core-fc",
        "name":     "Carolina Core FC",
        "short":    "CCFC",
        "aliases":  ["carolina core fc youth", "carolina core fc", "carolina core", "ccfc"],
        "sanityNameHints": ["carolina core", "ccfc", "core"],
        "label_strip":     ["carolina core fc youth", "carolina core fc", "carolina core", "ccfc"],
    },
    # New club that first appears in the girls schedule — placeholder name.
    # Rename `name` when we learn what NCC actually stands for and upload a
    # real crest to Sanity.
    {
        "id":       "ncc",
        "name":     "NCC",
        "short":    "NCC",
        "aliases":  ["ncc"],
        "sanityNameHints": [],
        "label_strip":     ["ncc"],
    },
]

# ─── Parsers ────────────────────────────────────────────────────────────────

WORD_RE = re.compile(r"\s+")

def normalise_ws(s: str) -> str:
    return WORD_RE.sub(" ", (s or "").strip())

# Age group: match U9/U10/…/U19 OR 9u/10u/… OR birth-year 2007–2018.
BIRTH_YEAR_TO_AGE_FALL_2026 = {
    2007: "U19", 2008: "U18", 2009: "U17", 2010: "U16",
    2011: "U15", 2012: "U14", 2013: "U13", 2014: "U12",
    2015: "U12", 2016: "U11", 2017: "U10", 2018: "U9",
}
AGE_RE = re.compile(
    r"(?:"
    r"\bU\s?(\d{1,2})(?![\d])"        # U11 / U 11 / U11M
    r"|\b(\d{1,2})\s?[uU](?![\w])"    # 11u / 11U
    r"|\b(20\d{2})\b"                 # 2015
    r"|^\s*(1[4-9])\b"                # bare 2-digit birth year at start ("15 CISC N Blue G")
    r")",
    re.I,
)

def parse_age(text: str) -> str | None:
    """Return "U11" / "U12" / etc. from any of: U11 · 11u · 2015. None if unclear."""
    if not text:
        return None
    m = AGE_RE.search(text)
    if not m:
        return None
    if m.group(1) or m.group(2):
        n = int(m.group(1) or m.group(2))
        return f"U{n}" if 8 <= n <= 19 else None
    if m.group(3):
        year = int(m.group(3))
        return BIRTH_YEAR_TO_AGE_FALL_2026.get(year)
    if m.group(4):
        # Bare 2-digit birth-year prefix (14–19) → 2014–2019.
        year = 2000 + int(m.group(4))
        return BIRTH_YEAR_TO_AGE_FALL_2026.get(year)
    return None

# Tier: A / B suffix if present (in team label or division).
TIER_RE = re.compile(r"\b([AB])\s*(?:team)?\s*$", re.I)

def parse_tier(*texts: str) -> str | None:
    for t in texts:
        if not t:
            continue
        m = TIER_RE.search(t.strip())
        if m:
            return m.group(1).upper()
    return None

# Gender: passed in per-feed since the tab is authoritative — girls tab is all G,
# boys tab is all M. Kept as a no-op function for future signal-based inference.
def parse_gender_from_division(division: str | None) -> str | None:
    if division and re.search(r"\bG\b", division):
        return "G"
    return None

def parse_date_loose(v):
    """Accept a datetime OR a string like '12.12.26' (MM.DD.YY) → date. None otherwise."""
    if isinstance(v, dt.datetime): return v.date()
    if isinstance(v, dt.date):     return v
    if not isinstance(v, str):     return None
    s = v.strip()
    m = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$", s)
    if not m: return None
    a, b, c = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if c < 100: c += 2000
    # Assume MM.DD.YY (US convention).
    try:
        return dt.date(c, a, b)
    except ValueError:
        return None

def parse_time_loose(v) -> tuple[int, int] | None:
    """Accept datetime.time OR strings like '10.30 AM' / '1.00 pm' / '13:30'."""
    if v is None or v == "":
        return None
    if isinstance(v, dt.time):
        return (v.hour, v.minute)
    if isinstance(v, dt.datetime):
        return (v.hour, v.minute)
    s = str(v).strip().lower()
    m = re.match(r"^\s*(\d{1,2})\s*[.:]\s*(\d{2})\s*(am|pm|a\.m\.|p\.m\.)?\s*$", s)
    if not m:
        # Bare hour: "9 AM"
        m2 = re.match(r"^\s*(\d{1,2})\s*(am|pm)\s*$", s)
        if not m2:
            return None
        hour = int(m2.group(1))
        minute = 0
        suffix = m2.group(2)
    else:
        hour   = int(m.group(1))
        minute = int(m.group(2))
        suffix = m.group(3)
    if suffix:
        suffix = suffix.replace(".", "")
        if suffix == "pm" and hour < 12:
            hour += 12
        elif suffix == "am" and hour == 12:
            hour = 0
    return (hour, minute)

def match_club(team_str: str) -> dict | None:
    """Return the canonical club dict matching this team string, or None."""
    if not team_str:
        return None
    s = normalise_ws(team_str).lower()
    best = None
    best_len = 0
    for club in CDL_CLUBS:
        for alias in club["aliases"]:
            if alias in s and len(alias) > best_len:
                best = club
                best_len = len(alias)
    return best

def derive_team_label(team_str: str, club: dict | None, age: str | None, tier: str | None) -> str:
    """Preserve the team-specific suffix (e.g. "CLT Man City 1", "South Red",
    "WS RED", "CDL 2"). Parents need this to find their child's actual team.

    Algorithm:
      1. Strip every age token from the raw string (U11 / 11u / 2015 / …).
      2. Strip the club's canonical label_strip tokens (longest first) so
         only the team-distinguishing suffix survives.
      3. Prepend the age, if known, so the label reads "U12 Man City 1".
    """
    s = normalise_ws(team_str)

    # 0. Trailing girls-suffix "G" — appears on nearly every girls team; redundant
    #    next to the gender filter, so strip it before further parsing.
    s = re.sub(r"\s+G$", "", s)

    # 1. Age tokens anywhere in the string.
    s = re.sub(
        r"(?:\bU\s?\d{1,2}[Mm]?\b|\b\d{1,2}[uU]\b|\b20\d{2}\b|^\s*1[4-9]\b)",
        "",
        s,
        flags=re.I,
    )

    # 2. Club tokens — longest-first so "carolina core fc youth" wins over "ccfc".
    if club:
        for token in sorted(club.get("label_strip", [club.get("short", "")]), key=len, reverse=True):
            if token:
                s = re.sub(r"\b" + re.escape(token) + r"\b", "", s, flags=re.I)

    # 3. Clean up leftover separators & whitespace.
    s = re.sub(r"[\s\-_/]+", " ", s).strip(" -–—.,")

    parts = []
    if age: parts.append(age)
    if s:   parts.append(s)
    if parts:
        return " ".join(parts)
    return tier or team_str.strip()

# ─── Placeholder crest (data-uri SVG per club, brand-tinted) ───────────────

CLUB_TINT = {
    "csa":                     ("#0B3D91", "#F2F2F2"),  # navy / white
    "cisc":                    ("#00589C", "#F2F2F2"),
    "nc-fusion":               ("#C8102E", "#F2F2F2"),
    "nc-fc-youth":             ("#B30838", "#F2F2F2"),
    "wilmington-hammerheads":  ("#003057", "#3EB1C8"),
    "highland-fc":             ("#5A5A5A", "#F2F2F2"),
    "sc-surf":                 ("#005A8B", "#F2F2F2"),
    "wake-fc":                 ("#003C71", "#F2F2F2"),
    "carolina-core-fc":        ("#0E6E4A", "#F2F2F2"),
    "ncc":                     ("#7C3AED", "#F2F2F2"),  # purple placeholder
    "unknown":                 ("#1E2D45", "#94A3B8"),
}

def placeholder_svg(club_id: str, short: str) -> str:
    bg, fg = CLUB_TINT.get(club_id, CLUB_TINT["unknown"])
    svg = (
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 88 88">'
        f'<circle cx="44" cy="44" r="42" fill="{bg}"/>'
        f'<text x="44" y="55" font-family="Barlow Condensed, Impact, sans-serif" '
        f'font-weight="900" font-size="30" fill="{fg}" text-anchor="middle" '
        f'letter-spacing="2">{short}</text></svg>'
    )
    import base64
    b64 = base64.b64encode(svg.encode()).decode()
    return f"data:image/svg+xml;base64,{b64}"

# ─── Main ──────────────────────────────────────────────────────────────────

def ingest_snapshot(snapshot_path: Path, gender: str,
                    used_clubs: dict, matches: list) -> int:
    """Load matches + clubs from a pre-baked snapshot JSON (used when the live
    source .xlsx has been removed from disk). Returns the number of matches added."""
    data = json.loads(snapshot_path.read_text())
    for c in data.get("clubs", []):
        used_clubs.setdefault(c["id"], {
            "id":              c["id"],
            "name":            c["name"],
            "short":           c.get("shortName") or c["id"][:3].upper(),
            "sanityNameHints": c.get("sanityNameHints", []),
        })
    added = 0
    for m in data.get("matches", []):
        matches.append({**m, "gender": gender})
        added += 1
    return added


def ingest_tab(path: Path, tab: str, gender: str, has_location: bool,
               used_clubs: dict, matches: list, warnings: list) -> tuple[int, int]:
    """Return (matches_added, rows_skipped) for a single (file, tab, gender)."""
    if not path.exists():
        print(f"  ⚠ Source not found, skipping: {path}")
        return (0, 0)
    wb = openpyxl.load_workbook(path, data_only=True)
    if tab not in wb.sheetnames:
        print(f"  ⚠ Tab '{tab}' not in {path.name}, skipping.")
        return (0, 0)
    ws = wb[tab]
    added = 0
    skipped = 0
    row_col_count = 7 if has_location else 6

    for r in range(2, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, row_col_count + 1)]
        date_v, time_v, field_v, home_v, away_v, div_v = cells[:6]
        loc_v = cells[6] if has_location else None

        if not date_v and not home_v and not away_v:
            continue  # blank row

        row_warnings: list[str] = []

        # Date
        date_only = parse_date_loose(date_v)
        if date_only is None:
            row_warnings.append(f"missing/invalid date: {date_v!r}")
            skipped += 1
            warnings.append({"gender": gender, "row": r, "warnings": row_warnings})
            continue

        # Time — TBD or unparseable → midnight + warn, but still emit the row.
        hm = parse_time_loose(time_v)
        if hm is None:
            if isinstance(time_v, str) and time_v.strip().upper() == "TBD":
                row_warnings.append("time TBD — showing at 12:00 AM")
            else:
                row_warnings.append(f"couldn't parse time: {time_v!r}")
            hm = (0, 0)
        hour, minute = hm

        kickoff = dt.datetime(date_only.year, date_only.month, date_only.day, hour, minute).isoformat()

        # Division = age group (+ optional tier suffix on boys, or "Premier/Championship League" on girls).
        div_str = normalise_ws(str(div_v) if div_v is not None else "")
        age = parse_age(div_str)
        tier_from_div = parse_tier(div_str)

        def parse_side(raw):
            raw = normalise_ws(str(raw or ""))
            # TBD teams stay unmatched — placeholder label, no club identity.
            if raw.upper() == "TBD":
                return raw, None, age, tier_from_div, "TBD"
            club = match_club(raw)
            side_age  = age or parse_age(raw)
            side_tier = tier_from_div or parse_tier(raw)
            label = derive_team_label(raw, club, side_age, side_tier)
            return raw, club, side_age, side_tier, label

        home_raw, home_club, home_age, home_tier, home_label = parse_side(home_v)
        away_raw, away_club, away_age, away_tier, away_label = parse_side(away_v)

        if not home_club and home_raw.upper() != "TBD":
            row_warnings.append(f"could not identify home club: {home_raw!r}")
        if not away_club and away_raw.upper() != "TBD":
            row_warnings.append(f"could not identify away club: {away_raw!r}")

        final_age = age or home_age or away_age
        if not final_age:
            row_warnings.append(f"no age group parseable (division={div_str!r}, home={home_raw!r})")
            final_age = "U11"

        for c in (home_club, away_club):
            if c and c["id"] not in used_clubs:
                used_clubs[c["id"]] = c
        # Only insert the "unknown" bucket for genuinely unresolved (non-TBD) sides.
        if (not home_club and home_raw.upper() != "TBD") or (not away_club and away_raw.upper() != "TBD"):
            used_clubs.setdefault("unknown", {"id": "unknown", "name": "Unknown Club", "short": "?"})

        def side_id(raw, club):
            if raw.upper() == "TBD": return "tbd"
            return (club or {"id": "unknown"})["id"]
        if home_raw.upper() == "TBD" or away_raw.upper() == "TBD":
            used_clubs.setdefault("tbd", {"id": "tbd", "name": "TBD", "short": "TBD"})

        matches.append({
            "id":              f"cdl-fall-{gender.lower()}-{r}",
            "kickoff":         kickoff,
            "homeClubId":      side_id(home_raw, home_club),
            "awayClubId":      side_id(away_raw, away_club),
            "homeTeamLabel":   home_label,
            "awayTeamLabel":   away_label,
            "field":           normalise_ws(str(field_v or "")),
            "locationAddress": normalise_ws(str(loc_v or "")) if has_location else "",
            "ageGroup":        final_age,
            "gender":          gender,
            "notes":           None,
            "sourceRow":       r,
        })
        added += 1

        if row_warnings:
            warnings.append({
                "gender": gender, "row": r, "date": str(date_only),
                "home": home_raw, "away": away_raw, "division": div_str,
                "warnings": row_warnings,
            })

    return (added, skipped)


def main() -> None:
    matches: list = []
    warnings: list = []
    used_clubs: dict = {}

    per_feed = []
    for feed in FEEDS:
        path = feed["path"]
        gender = feed["gender"]
        if not path.exists() and feed.get("snapshot_fallback") and feed["snapshot_fallback"].exists():
            snap = feed["snapshot_fallback"]
            print(f"→ Ingesting {gender} from SNAPSHOT {snap.name} (live source {path.name} missing)")
            added = ingest_snapshot(snap, gender, used_clubs, matches)
            per_feed.append((gender, added, 0))
            continue
        print(f"→ Ingesting {feed['tab']} ({gender}) from {path.name}")
        added, skipped = ingest_tab(
            path, feed["tab"], gender, feed["has_location"],
            used_clubs, matches, warnings,
        )
        per_feed.append((gender, added, skipped))

    # Emit clubs[] with placeholder crests + Sanity name hints so the server
    # component can swap in the real crest at request time.
    clubs_out = []
    for cid, meta in used_clubs.items():
        if cid == "tbd":
            # No crest — the loader/component treats this as "no crest available".
            continue
        short = meta.get("short") or cid[:3].upper()
        clubs_out.append({
            "id":               cid,
            "name":             meta["name"],
            "shortName":        short,
            "conference":       "",
            "logoUrl":          placeholder_svg(cid, short),
            "sanityNameHints":  meta.get("sanityNameHints", []),
        })
    clubs_out.sort(key=lambda c: c["name"])

    OUT_MATCHES.write_text(json.dumps({
        "generatedAt": dt.datetime.now().isoformat(timespec="seconds"),
        "sources":     [{"path": str(f["path"]), "tab": f["tab"], "gender": f["gender"]} for f in FEEDS],
        "clubs":       clubs_out,
        "matches":     matches,
    }, indent=2))

    OUT_WARNINGS.write_text(json.dumps(warnings, indent=2))

    per_feed_summary = ", ".join(f"{g}={a}" for g, a, _ in per_feed)
    total_skipped = sum(s for _, _, s in per_feed)
    print(f"\n✓ {len(matches)} matches ingested ({per_feed_summary}), "
          f"{total_skipped} skipped, {len(warnings)} rows with warnings")
    print(f"  → {OUT_MATCHES.relative_to(REPO_ROOT)}")
    print(f"  → {OUT_WARNINGS.relative_to(REPO_ROOT)} ({len(warnings)} entries)")
    print(f"  → clubs identified: {[c['id'] for c in clubs_out]}")


if __name__ == "__main__":
    main()
